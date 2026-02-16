"""Testes para utilitários e validações.

Testa validações de Excel, exportação PDF e outros utilitários.
"""

import pytest
from unittest.mock import Mock, patch, MagicMock
import tempfile
import os

from src.common.validators.exel_validator import ExcelValidator
from src.common.utils.pdf_export import PDFExporter
from src.common.utils.notification import NotificationManager


class TestExcelValidator:
    """Testes para ExcelValidator."""

    def test_validar_arquivo_excel_valido(self):
        """Testa validação de arquivo Excel válido."""
        # Cria um arquivo Excel temporário real
        import openpyxl
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            temp_path = temp_file.name

        try:
            # Cria um workbook com alguns dados
            wb = openpyxl.Workbook()
            ws = wb.active
            for i in range(1, 11):
                ws[f"A{i}"] = f"Dado {i}"
            wb.save(temp_path)
            wb.close()

            validator = ExcelValidator()
            resultado = validator.validar_arquivo(temp_path)

            assert resultado["valido"] is True
            assert resultado["linhas"] > 0
        finally:
            os.unlink(temp_path)

    def test_validar_arquivo_excel_invalido(self):
        """Testa validação de arquivo Excel inválido."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, mode="w") as temp_file:
            temp_file.write("Arquivo inválido")
            temp_path = temp_file.name

        try:
            validator = ExcelValidator()
            resultado = validator.validar_arquivo(temp_path)

            assert resultado["valido"] is False
            assert len(resultado["erros"]) > 0
        finally:
            os.unlink(temp_path)

    def test_validar_arquivo_nao_existente(self):
        """Testa validação de arquivo não existente."""
        validator = ExcelValidator()
        resultado = validator.validar_arquivo("arquivo_nao_existente.xlsx")

        assert resultado["valido"] is False
        assert len(resultado["erros"]) > 0
        assert "Ficheiro não encontrado" in resultado["erros"][0]

    def test_validar_extensao_invalida(self):
        """Testa validação de arquivo com extensão inválida."""
        with tempfile.NamedTemporaryFile(suffix=".txt", delete=False, mode="w") as temp_file:
            temp_file.write("Conteúdo de teste")
            temp_path = temp_file.name

        try:
            validator = ExcelValidator()
            resultado = validator.validar_arquivo(temp_path)

            assert resultado["valido"] is False
            assert len(resultado["erros"]) > 0
            assert "Extensão não permitida" in resultado["erros"][0]
        finally:
            os.unlink(temp_path)


class TestPDFExporter:
    """Testes para PDFExporter."""

    def test_exportar_tabela_taxas_pdf(self):
        """Testa exportação de tabela de taxas para PDF."""
        # Cria dados de teste
        tabela_data = {
            "mes": 1,
            "ano": 2024,
            "total_dias_trabalho": 20,
            "total_km": 500.0,
            "total_ips": 160,
            "total_faturacao": 2500.0,
            "linhas": [
                {
                    "dia": 1,
                    "dia_semana": "Segunda-feira",
                    "tipo": "trabalho",
                    "ips": 8,
                    "valor_com_iva": 100.0,
                    "km": 25.0,
                    "locais": "Lisboa",
                }
            ]
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, "tabela_taxas.pdf")

            exporter = PDFExporter()
            resultado = exporter.exportar_mapa_taxas(tabela_data, output_path)

            assert resultado["sucesso"] is True
            assert os.path.exists(output_path)
            assert resultado["caminho"] == output_path

    def test_exportar_mapa_taxas_pdf(self):
        """Testa exportação de mapa de taxas para PDF."""
        # Cria dados de teste
        mapa_data = {
            "mes": 1,
            "ano": 2024,
            "total_setubal": 1000.0,
            "total_santarem": 500.0,
            "total_evora": 250.0,
            "total_ips": 80,
            "total_dividas": 100.0,
            "total_faturacao": 1750.0,
            "linhas": [
                {
                    "ordem": 1,
                    "recibo": 100,
                    "setubal": 500.0,
                    "santarem": 250.0,
                    "evora": 125.0,
                    "ips": 40,
                    "dividas": 50.0,
                }
            ]
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, "mapa_taxas.pdf")

            exporter = PDFExporter()
            resultado = exporter.exportar_mapa_taxas(mapa_data, output_path)

            assert resultado["sucesso"] is True
            assert os.path.exists(output_path)
            assert resultado["caminho"] == output_path

    def test_exportar_pdf_erro_arquivo(self):
        """Testa erro ao exportar PDF para caminho inválido."""
        tabela_data = {"mes": 1, "ano": 2024}

        # Tenta exportar para um caminho inválido
        exporter = PDFExporter()
        resultado = exporter.exportar_mapa_taxas(tabela_data, "/invalid/path/file.pdf")

        assert resultado["sucesso"] is False
        assert len(resultado["erro"]) > 0


class TestNotificationManager:
    """Testes para NotificationManager."""

    def test_enviar_notificacao_sucesso(self):
        """Testa envio de notificação de sucesso."""
        manager = NotificationManager()
        with patch.object(manager.logger, 'info') as mock_info:
            manager.enviar_sucesso("Operação concluída com sucesso")

            mock_info.assert_called_once_with("[SUCESSO] Operação concluída com sucesso")

    def test_enviar_notificacao_erro(self):
        """Testa envio de notificação de erro."""
        manager = NotificationManager()
        with patch.object(manager.logger, 'error') as mock_error:
            manager.enviar_erro("Ocorreu um erro na operação")

            mock_error.assert_called_once_with("[ERRO] Ocorreu um erro na operação")

    def test_enviar_notificacao_aviso(self):
        """Testa envio de notificação de aviso."""
        manager = NotificationManager()
        with patch.object(manager.logger, 'warning') as mock_warning:
            manager.enviar_aviso("Atenção: operação incompleta")

            mock_warning.assert_called_once_with("[AVISO] Atenção: operação incompleta")

    def test_enviar_notificacao_debug(self):
        """Testa envio de notificação de debug."""
        manager = NotificationManager()
        with patch.object(manager.logger, 'debug') as mock_debug:
            manager.enviar_debug("Detalhes da operação")

            mock_debug.assert_called_once_with("[DEBUG] Detalhes da operação")

    def test_formatar_mensagem(self):
        """Testa formatação de mensagens."""
        manager = NotificationManager()

        mensagem = manager.formatar_mensagem("SUCESSO", "Operação concluída")
        assert "SUCESSO" in mensagem
        assert "Operação concluída" in mensagem

        mensagem = manager.formatar_mensagem("ERRO", "Operação falhou")
        assert "ERRO" in mensagem
        assert "Operação falhou" in mensagem

    def test_notificacao_com_dados(self):
        """Testa notificação com dados adicionais."""
        manager = NotificationManager()
        with patch.object(manager.logger, 'info') as mock_info:
            dados = {"arquivo": "test.xlsx", "linhas": 100}
            manager.enviar_sucesso("Arquivo processado", dados)

            # Verifica se os dados foram incluídos na mensagem
            args = mock_info.call_args[0][0]
            assert "test.xlsx" in args
            assert "100" in args


class TestUtilsIntegracao:
    """Testes de integração para utilitários."""

    def test_fluxo_completo_validacao_exportacao(self):
        """Testa fluxo completo: validação Excel → exportação PDF."""
        import openpyxl
        # 1. Validação do Excel - Cria um arquivo Excel temporário real
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            temp_path = temp_file.name

        try:
            # Cria um workbook com alguns dados
            wb = openpyxl.Workbook()
            ws = wb.active
            for i in range(1, 11):
                ws[f"A{i}"] = f"Dado {i}"
            wb.save(temp_path)
            wb.close()

            validator = ExcelValidator()
            resultado_validacao = validator.validar_arquivo(temp_path)

            assert resultado_validacao["valido"] is True
        finally:
            os.unlink(temp_path)

        # 2. Exportação para PDF
        tabela_data = {
            "mes": 1,
            "ano": 2024,
            "total_dias_trabalho": 10,
            "linhas": []
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, "tabela.pdf")

            exporter = PDFExporter()
            resultado_exportacao = exporter.exportar_mapa_taxas(tabela_data, output_path)

            assert resultado_exportacao["sucesso"] is True
            assert os.path.exists(output_path)

    def test_notificacoes_em_cascata(self):
        """Testa notificações em cascata para diferentes operações."""
        manager = NotificationManager()
        with patch.object(manager.logger, 'info') as mock_info, \
             patch.object(manager.logger, 'warning') as mock_warning:

            # Simula uma sequência de operações
            manager.enviar_sucesso("Início da operação")
            manager.enviar_aviso("Processando dados")
            manager.enviar_sucesso("Operação concluída")

            # Verifica todas as chamadas
            assert mock_info.call_count == 2  # início e conclusão
            assert mock_warning.call_count == 1  # aviso

    def test_erro_em_validacao_excel(self):
        """Testa tratamento de erro em validação Excel."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, mode="w") as temp_file:
            temp_file.write("Arquivo inválido")
            temp_path = temp_file.name

        try:
            validator = ExcelValidator()
            resultado = validator.validar_arquivo(temp_path)

            # Verifica se o resultado é inválido e contém erros
            assert resultado["valido"] is False
            assert len(resultado["erros"]) > 0
            assert "Excel válido" in resultado["erros"][0]
        finally:
            os.unlink(temp_path)
